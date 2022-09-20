from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import os
from selenium.webdriver.common.action_chains import ActionChains
import time
from pathlib import Path
import sqlite3
import openpyxl
import win32com.client as win32
import pandas as pd
import string 
import patoolib
from rarfile import RarFile
def latest_download_file(folder_of_download):#Вытаскивает последний скачанный файл
      os.chdir(folder_of_download)
      files = sorted(os.listdir(os.getcwd()), key=os.path.getmtime)
      newest = files[-1]
      return newest

desktop = os.path.join(os.path.join(os.environ['USERPROFILE']), 'Desktop')
Path(desktop+"\\bts").mkdir(parents=True, exist_ok=True)
Path(desktop+"\\bts\\Download").mkdir(parents=True, exist_ok=True)
conn = sqlite3.connect(os.path.abspath(desktop+"\\bts\\btsdb.db"))
cursor = conn.cursor()
orders_q = """            
create table if not exists Status (
	"№"	INTEGER,
	"БИН/ИИН должника"	TEXT,
	"Наименование Ф.И.О.должника"	TEXT,
	"Номер государственной регистрации должника"	TEXT,
	"Адрес местонахождения должника"	TEXT,
	"Наименование суда"	TEXT,
	"Дата вынесения определения о возбуждении дела о банкротстве"	TEXT,
	"Дата назначения временного управляющего"	TEXT,
	"Ф.И.О. Временного управляющего"	TEXT,
	"Срок принятия требований кредиторов временным управляющим С"	TEXT,
	"Срок принятия требований кредиторов временным управляющим До"	TEXT,
	"Адрес приема требований"	TEXT,
	"Контактные данные (телефон, электронный адрес) временного управляющего"	TEXT,
	"Дата размещения объявления"	TEXT,
	PRIMARY KEY("№" AUTOINCREMENT)
);
"""
cursor.execute(orders_q)

alphabet = string.ascii_uppercase 
chrome_options = webdriver.ChromeOptions()
prefs = {
"download.default_directory": desktop+"\\bts\\Download",
"download.directory_upgrade": True
}
chrome_options.add_experimental_option('prefs', prefs)
driver = webdriver.Chrome(ChromeDriverManager().install(),chrome_options=chrome_options)


def parse_url(url):#парсит сайт
    print(f"{url}")
    driver.get(f"{url}")
    time.sleep(5)
    #driver.find_elements_by_xpath("//a[text()='Физическим лицам']")
    Par_element=driver.find_elements_by_tag_name("a")
    for el in Par_element:
        #print(el.text.lower())
        if "юридическим лицам" in el.text.lower():
            el.click()
            break
    element=driver.find_element_by_xpath("// a[contains(text(),'Реабилитация и банкротство')]")
    action = ActionChains(driver)
    action.double_click(on_element = element)
    action.perform()
    time.sleep(5)
    element=driver.find_element_by_class_name("catmenu").text
    #driver.execute_script("window.history.go(-1)")#back
    element=element.split("\n")
    years=[]
    for e in element:
        if e not in years and len(e)==8:
            years.append(e)
    downloaded_files=[]        
    count=0
    for year in years:
        count+=1
        if count==9:
            break
        parentElement=driver.find_element_by_class_name("menu")
        elementList = parentElement.find_elements_by_xpath("// a[contains(text(),'%s')]"%year)
        print(year)
        for ele in elementList:
            print( ele.get_attribute('href'))
            if ele.get_attribute('href').endswith('god'):
                result = None
                href=ele.get_attribute('href')
                while result is None:
                    try:
                        driver.get(href)
                        time.sleep(5)
                        #print(1,result)
                        inf_el=driver.find_elements_by_xpath('//a[contains(text(),"Информаци")]')
                        result=1
                    except:
                         pass
                for infmes in inf_el:
                    if infmes.text=="Информационные сообщения" or infmes.text=="Информационное сообщение":
                        result = None
                        href=infmes.get_attribute('href')
                        while result is None:
                            try:
                                driver.get(href)
                                time.sleep(5)
                                #print(2,result)
                                parentElement_download=driver.find_elements_by_tag_name("a")
                                result=1
                            except:
                                pass
                        for el in parentElement_download:
                            if el.text.endswith('кредиторами временному управляющему') or el.text.endswith('кредиторами временному управляющему.'):
                                #print(el.text)
				typee=1 if el.text.endswith('кредиторами временному управляющему') else 0
                                result = None
                                href=el.get_attribute('href')
                                while result is None:
                                    try:
                                        driver.get(href)
                                        time.sleep(5)
                                        el.text.endswith('кредиторами временному управляющему') if typee==1 else el.text.endswith('кредиторами временному управляющему.')
                                        #print(3,result)
                                        result=1
                                    except:
                                        try:
                                            driver.find_element_by_xpath('//*[contains(text(),"Not Found")]')
                                            driver.execute_script("window.history.go(-1)")
                                            #print(4,result)
                                            result=1
                                        except:
                                            pass
                        driver.execute_script("window.history.go(-1)")
                        downloaded_files.append(latest_download_file(desktop+"\\bts\\Download"))
                        break
                time.sleep(2)
                driver.execute_script("window.history.go(-1)")
                    #ws.cell.get_column_letter(ws.max_row)
                break
        print("END")
    downloaded_files = list(dict.fromkeys(downloaded_files))
    return downloaded_files


def read_insert(downloaded_files):#читает и записывает в sqlite
    for file in downloaded_files:
        try:
            excel_file = openpyxl.load_workbook(desktop+"\\bts\\Download\\%s"%file)
        except:
            if file.endswith("xls"):
                try:
                    fname = os.path.abspath(desktop+"\\bts\\Download\\%s"%file) 
                    excel = win32.gencache.EnsureDispatch('Excel.Application')
                    #excel.Visible = True
                    wb = excel.Workbooks.Open( os.path.join(desktop+"\\bts\\Download\\", file), CorruptLoad=1)
                    wb.SaveAs(fname+"x", FileFormat = 51)    
                    wb.Close()                               
                    excel.Application.Quit()
                except:
                    df = pd.read_excel(desktop+"\\bts\\Download\\%s"%file, header=None)
                    df.to_excel(desktop+"\\bts\\Download\\%s"%file+"x", index=False, header=False)
                os.remove(desktop+"\\bts\\Download\\%s"%file)
                excel_file = openpyxl.load_workbook(desktop+"\\bts\\Download\\%s"%file+'x')
            elif file.endswith("rar"):
                try:
                    old_file=file
                    with RarFile(desktop+"\\bts\\Download\\%s"%file) as file:
                        new_file=file.namelist()[0]
                    patoolib.extract_archive(desktop+"\\bts\\Download\\%s"%old_file, outdir=desktop+"\\bts\\Download")
                    os.remove(desktop+"\\bts\\Download\\%s"%old_file)
                    excel_file = openpyxl.load_workbook(desktop+"\\bts\\Download\\%s"%new_file)
                except:
                    continue
        ws=excel_file.active
        print(ws.max_column,file)
        add_to_db=False
        for x in range(1,ws.max_row+1):
            if type(ws[f"B{x}"].value)!=type(None) and "временному управляющему" not in str(ws[f"B{x}"].value):
                header_row=x
                break
        last_column_letter= 0
        column_names=[]
        for x in range(0,len(alphabet)):
            #print(ws[f"{alphabet[x]}{header_row}"].value)
            if ws[f"{alphabet[x]}{header_row}"].value is not None:
                column_names.append(ws[f"{alphabet[x]}{header_row}"].value)
                last_column_letter+=1
                
        alphabet_names={}
        letter=iter(list(alphabet))
        next(letter)
        for x in range(len(column_names)):
            if "номер должника" in column_names[x] or "БИН/ИИН" in column_names[x]:
                alphabet_names.update({'B': next(letter)})
            elif "наименование должника" in column_names[x] or "Ф.И.О.должника" in column_names[x]:
                alphabet_names.update({'C': next(letter)})
            elif "регистрации должника" in column_names[x]:
                alphabet_names.update({'D': next(letter)})
            elif "местонахождения должника" in column_names[x]:
                alphabet_names.update({'E': next(letter)})
            elif "наименование суда" in column_names[x].lower():
                alphabet_names.update({'F': next(letter)})
            elif "определения о возбуждении" in column_names[x]:
                alphabet_names.update({'G': next(letter)})
            elif "назначения временного управляющего" in column_names[x] or "полномочий временного управлющего" in column_names[x] or "дата приказа" in column_names[x]:
                alphabet_names.update({'H': next(letter)})
            elif "Ф.И.О. Временного управляющего" in column_names[x] or "личность) временного управляющего" in column_names[x]:
                alphabet_names.update({'I': next(letter)})
            elif "Срок принятия" in column_names[x]:
                alphabet_names.update({'J': next(letter)})
                alphabet_names.update({'K': next(letter)})
            elif "Адрес приема" in column_names[x]:
                alphabet_names.update({'L': next(letter)})
            elif "Контактные данные" in column_names[x]:
                alphabet_names.update({'M': next(letter)})
            elif "Дата размещения" in column_names[x]:
                alphabet_names.update({'N': next(letter)})
            else:
                if '№' not in column_names[x]:
                    next(letter)
        for x in list(alphabet)[1:14]:
            try:
                print(alphabet_names[x])
            except:
                alphabet_names.update({x: None})
                print(x+"-NE")
        for x in range(1,ws.max_row+1):
            if add_to_db==True:
                if ws[f"B{x}"].value is None and ws[f"C{x}"].value is None and ws[f"D{x}"].value is None and ws[f"E{x}"].value is None and ws[f"D{x}"].value is None:
                    break
                #print(ws[f"A{x}"].value)
                orders_q = """INSERT INTO Status (
                    "БИН/ИИН должника", "Наименование Ф.И.О.должника", "Номер государственной регистрации должника", 
                    "Адрес местонахождения должника", "Наименование суда", "Дата вынесения определения о возбуждении дела о банкротстве",
                    "Дата назначения временного управляющего", "Ф.И.О. Временного управляющего",
                    "Срок принятия требований кредиторов временным управляющим С", "Срок принятия требований кредиторов временным управляющим До",
                    "Адрес приема требований", "Контактные данные (телефон, электронный адрес) временного управляющего", "Дата размещения объявления" 
                    ) SELECT '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s', '%s'
                    WHERE NOT EXISTS (SELECT 1 FROM Status WHERE "БИН/ИИН должника" = '%s' AND
                    "Срок принятия требований кредиторов временным управляющим С" = '%s');"""%(ws[f"{alphabet_names['B']}{x}"].value if alphabet_names['B'] is not None else None,(ws[f"{alphabet_names['C']}{x}"].value.replace( '"', '').replace( "'", '')  if type(ws[f"{alphabet_names['C']}{x}"].value)==str else int(str(ws[f"{alphabet_names['C']}{x}"].value).replace( '"', '').replace( "'", ''))) if alphabet_names['C'] is not None else None,
                    ws[f"{alphabet_names['D']}{x}"].value if alphabet_names['D'] is not None else None,str(ws[f"{alphabet_names['E']}{x}"].value).replace( '\n', '') if alphabet_names['E'] is not None else None,ws[f"{alphabet_names['F']}{x}"].value if alphabet_names['F'] is not None else None,ws[f"{alphabet_names['G']}{x}"].value if alphabet_names['G'] is not None else None,ws[f"{alphabet_names['H']}{x}"].value if alphabet_names['H'] is not None else None,
                    ws[f"{alphabet_names['I']}{x}"].value if alphabet_names['I'] is not None else None,ws[f"{alphabet_names['J']}{x}"].value if alphabet_names['J'] is not None else None,ws[f"{alphabet_names['K']}{x}"].value if alphabet_names['K'] is not None else None,ws[f"{alphabet_names['L']}{x}"].value if alphabet_names['L'] is not None else None,ws[f"{alphabet_names['M']}{x}"].value if alphabet_names['M'] is not None else None,ws[f"{alphabet_names['N']}{x}"].value if alphabet_names['N'] is not None else None,
                    ws[f"{alphabet_names['B']}{x}"].value if alphabet_names['B'] is not None else None,ws[f"{alphabet_names['J']}{x}"].value if alphabet_names['J'] is not None else None)
                cursor.execute(orders_q)
                conn.commit()
            if ws[f"A{x}"].value==1:
                add_to_db=True


if __name__ == "__main__":
    urls=["http://vko.kgd.gov.kz/ru/", "http://zhmb.kgd.gov.kz/ru/", "http://zko.kgd.gov.kz/ru/", 
          "http://krg.kgd.gov.kz/ru/", "http://kst.kgd.gov.kz/ru/", "http://kzl.kgd.gov.kz/ru/", 
          "http://mng.kgd.gov.kz/ru/", "http://pvl.kgd.gov.kz/ru/", "http://sko.kgd.gov.kz/ru/",
          "http://trk.kgd.gov.kz/ru/", "http://nursultan.kgd.gov.kz/ru/" , "http://almaty.kgd.gov.kz/ru/", 
          "http://shymkent.kgd.gov.kz/ru/", "http://akb.kgd.gov.kz/ru/", "http://alm.kgd.gov.kz/ru/", 
          "http://atr.kgd.gov.kz/ru/",  "http://akm.kgd.gov.kz/ru/"]
    for url in urls:
        read_insert(parse_url(url))

conn.close() 
driver.quit()
